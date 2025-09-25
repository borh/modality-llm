"""
Functions for loading, generating, and managing examples.
"""

import asyncio
import hashlib
import json
import logging
import os
import re
from pathlib import Path
from typing import List, Literal, Optional

import requests
from pydantic import BaseModel
from tqdm import tqdm

from modality_llm.schema import Example, GrammarLabel
from modality_llm.utils import load_jsonl_models, mark_changed_tokens


def determine_transformation_strategy_from_augmentation(
    entry: "Example", alt_modal: str
) -> str:
    """
    Determine transformation strategy using semantic relationships between modals.
    """
    original_modal = (entry.english_target or "").lower()
    alt_modal_lower = alt_modal.lower()

    # First try the existing augmentation system
    from modality_llm.augmentation import (
        generate_contradiction_variants,
        generate_entailment_tests,
        generate_substitution_variants,
    )

    # Check substitution variants
    try:
        substitution_variants = generate_substitution_variants(entry, max_variants=10)
        for variant in substitution_variants:
            if (
                variant.english_target
                and variant.english_target.lower() == alt_modal_lower
            ):
                return variant.transformation_strategy
    except Exception:
        pass

    # Check entailment tests
    try:
        entailment_tests = generate_entailment_tests(entry, max_hypotheses=10)
        for test in entailment_tests:
            if alt_modal_lower in test.hypothesis.lower():
                return test.transformation_strategy
    except Exception:
        pass

    # Check contradiction variants
    try:
        contradiction_variants = generate_contradiction_variants(entry, max_variants=10)
        for variant in contradiction_variants:
            if alt_modal_lower in variant.hypothesis.lower():
                return variant.transformation_strategy
    except Exception:
        pass

    # Fallback to semantic analysis
    return analyze_modal_relationship(original_modal, alt_modal_lower)


def analyze_modal_relationship(original: str, alternative: str) -> str:
    """
    Analyze semantic relationship between two modals to determine transformation strategy.
    """
    # Define semantic relationships between modals
    modal_semantics = {
        "can": {
            "category": "ability",
            "strength": "medium",
            "polarity": "positive",
        },
        "could": {
            "category": "ability",
            "strength": "weak",
            "polarity": "positive",
        },
        "cannot": {
            "category": "ability",
            "strength": "strong",
            "polarity": "negative",
        },
        "may": {
            "category": "permission",
            "strength": "medium",
            "polarity": "positive",
        },
        "might": {
            "category": "possibility",
            "strength": "weak",
            "polarity": "positive",
        },
        "must": {
            "category": "necessity",
            "strength": "strong",
            "polarity": "positive",
        },
        "should": {
            "category": "advice",
            "strength": "medium",
            "polarity": "positive",
        },
        "will": {
            "category": "prediction",
            "strength": "strong",
            "polarity": "positive",
        },
        "would": {
            "category": "conditional",
            "strength": "medium",
            "polarity": "positive",
        },
        "shall": {
            "category": "obligation",
            "strength": "strong",
            "polarity": "positive",
        },
        "ought to": {
            "category": "advice",
            "strength": "medium",
            "polarity": "positive",
        },
    }

    orig_sem = modal_semantics.get(original, {})
    alt_sem = modal_semantics.get(alternative, {})

    if not orig_sem or not alt_sem:
        return f"modal_substitution_{original}_to_{alternative}"

    orig_cat = orig_sem.get("category", "unknown")
    alt_cat = alt_sem.get("category", "unknown")
    orig_strength = orig_sem.get("strength", "medium")
    alt_strength = alt_sem.get("strength", "medium")
    orig_pol = orig_sem.get("polarity", "positive")
    alt_pol = alt_sem.get("polarity", "positive")

    # Same category transformations
    if orig_cat == alt_cat:
        if orig_strength == alt_strength:
            return f"{orig_cat}_paraphrase"
        elif strength_order(alt_strength) > strength_order(orig_strength):
            return f"{orig_cat}_strengthening"
        else:
            return f"{orig_cat}_weakening"

    # Polarity changes
    if orig_pol != alt_pol:
        return f"{orig_cat}_to_negation"

    # Cross-category transformations (entailment relationships)
    cross_category_patterns = {
        ("necessity", "advice"): "necessity_to_advice",
        ("necessity", "permission"): "necessity_to_permission",
        ("necessity", "possibility"): "necessity_to_possibility",
        ("advice", "permission"): "advice_to_permission",
        ("advice", "possibility"): "advice_to_possibility",
        ("ability", "permission"): "ability_to_permission",
        ("ability", "possibility"): "ability_to_possibility",
        ("prediction", "possibility"): "prediction_to_possibility",
        ("obligation", "advice"): "obligation_to_advice",
        ("conditional", "possibility"): "conditional_to_possibility",
    }

    pattern_key = (orig_cat, alt_cat)
    if pattern_key in cross_category_patterns:
        return cross_category_patterns[pattern_key]

    # Reverse entailment (contradiction)
    reverse_patterns = {
        ("advice", "necessity"): "advice_to_necessity",
        ("permission", "necessity"): "permission_to_necessity",
        ("possibility", "necessity"): "possibility_to_necessity",
        ("possibility", "prediction"): "possibility_to_prediction",
    }

    if pattern_key in reverse_patterns:
        return reverse_patterns[pattern_key]

    # Default fallback with category information
    return f"{orig_cat}_to_{alt_cat}_substitution"


def strength_order(strength: str) -> int:
    """Convert strength to numeric order for comparison."""
    return {"weak": 1, "medium": 2, "strong": 3}.get(strength, 2)


def get_transformation_category(strategy: str) -> str:
    """Determine the high-level category of transformation."""
    if (
        "paraphrase" in strategy
        or "strengthening" in strategy
        or "weakening" in strategy
    ):
        return "substitution"
    elif "to_negation" in strategy or "to_denial" in strategy:
        return "contradiction"
    elif "_to_" in strategy and "substitution" not in strategy:
        return "entailment"
    elif strategy == "original":
        return "original"
    else:
        return "substitution"  # default


def download_modal_verb_dataset(target_path="modal_verbs.jsonl") -> str:
    path = Path(target_path)
    if path.exists():
        print(f"Dataset already exists at {path}")
        return str(path)

    url = "https://raw.githubusercontent.com/minnesotanlp/moverb/main/data/moVerb_all.jsonl"
    print(f"Downloading dataset from {url}...")

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(response.content)
        print(f"Dataset successfully downloaded to {path}")
        return str(path)
    except Exception as e:
        print(f"Error: {e}")
        # Ensure parent directories exist before creating file
        path.parent.mkdir(parents=True, exist_ok=True)
        path.touch()
        return str(path)


def generate_grammar_examples_xlsx(
    csv_rows: list[dict],
    output_xlsx_path: str,
) -> None:
    """Generate XLSX file with dropdown validation and formatting for annotation, including conditional formatting for completion."""
    try:
        import xlsxwriter
    except ImportError:
        print("Error: xlsxwriter not installed. Install with: pip install xlsxwriter")
        return

    output_path = Path(output_xlsx_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook and worksheet
    workbook = xlsxwriter.Workbook(str(output_path))
    worksheet = workbook.add_worksheet("Annotation")

    # Define formats
    header_format = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#D7E4BC",
            "border": 1,
            "text_wrap": True,
            "valign": "vcenter",
        }
    )

    # Format for cells that should wrap (only Marked_Sentence_English)
    cell_format_wrap = workbook.add_format(
        {"border": 1, "text_wrap": True, "valign": "top"}
    )

    # Format for cells that should NOT wrap (single row height)
    cell_format_nowrap = workbook.add_format(
        {"border": 1, "text_wrap": False, "valign": "vcenter"}
    )

    # Format for completed rows (light green background)
    completed_format = workbook.add_format(
        {
            "border": 1,
            "text_wrap": False,
            "valign": "vcenter",
            "bg_color": "#E8F5E8",  # Light green background
        }
    )

    # Format for completed rows with text wrapping
    completed_format_wrap = workbook.add_format(
        {
            "border": 1,
            "text_wrap": True,
            "valign": "top",
            "bg_color": "#E8F5E8",  # Light green background
        }
    )

    if not csv_rows:
        workbook.close()
        return

    # Get column names and find the Completed column index
    columns = list(csv_rows[0].keys())
    completed_col_idx = None
    try:
        completed_col_idx = columns.index("Completed")
    except ValueError:
        print(
            "Warning: 'Completed' column not found. Conditional formatting will not be applied."
        )

    # Define dropdown options
    palmer_options = ["deontic", "epistemic", "dynamic", "unknown"]
    quirk_options = [
        "possibility",
        "ability",
        "permission",
        "necessity",
        "obligation",
        "inference",
        "prediction",
        "volition",
        "unknown",
    ]
    grammaticality_options = ["yes", "no", "partial"]  # Add "partial"
    completed_options = ["0", "1"]

    # Hidden columns (will be hidden completely)
    hidden_columns = [
        "ID",
        "EID",
        "Original_Sentence",
        "Modal_Verb",
        "Marked_Sentence_Japanese",
        "Palmer_Annotations",
        "Quirk_Annotations",
    ]

    # Write headers and set column properties
    for col_idx, col_name in enumerate(columns):
        worksheet.write(0, col_idx, col_name, header_format)

        # Convert column index to Excel column letter (supports up to ZZ)
        if col_idx < 26:
            col_letter = chr(65 + col_idx)
        else:
            col_letter = chr(65 + (col_idx // 26) - 1) + chr(65 + (col_idx % 26))

        # Set column widths and hide columns
        if col_name in hidden_columns:
            worksheet.set_column(
                f"{col_letter}:{col_letter}",
                20,
                cell_format_nowrap,
                {"hidden": 1},
            )
        elif col_name == "Marked_Sentence_English":
            # Only this column should wrap text
            worksheet.set_column(f"{col_letter}:{col_letter}", 40, cell_format_wrap)
        elif col_name == "Annotation_Notes":
            # Wide column but no wrapping - let it extend horizontally
            worksheet.set_column(f"{col_letter}:{col_letter}", 60, cell_format_nowrap)
        elif col_name in [
            "Palmer_Expected",
            "Quirk_Expected",
            "Grammaticality_Expected",
            "Completed",
        ]:
            worksheet.set_column(f"{col_letter}:{col_letter}", 15, cell_format_nowrap)
        elif col_name in ["Transformation_Strategy", "Transformation_Category"]:
            worksheet.set_column(f"{col_letter}:{col_letter}", 20, cell_format_nowrap)
        else:
            worksheet.set_column(f"{col_letter}:{col_letter}", 12, cell_format_nowrap)

    # Write data and add validation
    for row_idx, row_data in enumerate(csv_rows, start=1):
        # Check if this row is completed
        is_completed = str(row_data.get("Completed", "0")) == "1"

        for col_idx, col_name in enumerate(columns):
            value = row_data.get(col_name, "")

            # Choose format based on column type and completion status
            if col_name in hidden_columns:
                fmt = completed_format if is_completed else cell_format_nowrap
            elif col_name == "Marked_Sentence_English":
                fmt = completed_format_wrap if is_completed else cell_format_wrap
            else:
                fmt = completed_format if is_completed else cell_format_nowrap

            worksheet.write(row_idx, col_idx, value, fmt)

            # Add dropdown validation for specific columns
            if col_name == "Palmer_Expected":
                worksheet.data_validation(
                    row_idx,
                    col_idx,
                    row_idx,
                    col_idx,
                    {
                        "validate": "list",
                        "source": palmer_options,
                        "dropdown": True,
                        "error_message": "Please select a valid Palmer category.",
                    },
                )
            elif col_name == "Quirk_Expected":
                worksheet.data_validation(
                    row_idx,
                    col_idx,
                    row_idx,
                    col_idx,
                    {
                        "validate": "list",
                        "source": quirk_options,
                        "dropdown": True,
                        "error_message": "Please select a valid Quirk category.",
                    },
                )
            elif col_name == "Grammaticality_Expected":
                worksheet.data_validation(
                    row_idx,
                    col_idx,
                    row_idx,
                    col_idx,
                    {
                        "validate": "list",
                        "source": grammaticality_options,
                        "dropdown": True,
                        "error_message": "Please select yes or no.",
                    },
                )
            elif col_name == "Completed":
                worksheet.data_validation(
                    row_idx,
                    col_idx,
                    row_idx,
                    col_idx,
                    {
                        "validate": "list",
                        "source": completed_options,
                        "dropdown": True,
                        "error_message": "Please select 0 (incomplete) or 1 (complete).",
                    },
                )

    # Add conditional formatting for dynamic highlighting when Completed changes
    if completed_col_idx is not None:
        # Convert completed column index to Excel column letter
        if completed_col_idx < 26:
            completed_col_letter = chr(65 + completed_col_idx)
        else:
            completed_col_letter = chr(65 + (completed_col_idx // 26) - 1) + chr(
                65 + (completed_col_idx % 26)
            )

        # Apply conditional formatting to the entire data range
        last_row = len(csv_rows)
        last_col_letter = (
            chr(65 + len(columns) - 1)
            if len(columns) <= 26
            else chr(65 + (len(columns) - 1) // 26 - 1)
            + chr(65 + (len(columns) - 1) % 26)
        )

        # Format for when Completed = 1 (light green)
        worksheet.conditional_format(
            f"A2:{last_col_letter}{last_row + 1}",  # Data range (excluding header)
            {
                "type": "formula",
                "criteria": f"=${completed_col_letter}2=1",  # Check if Completed column = 1
                "format": workbook.add_format({"bg_color": "#E8F5E8"}),  # Light green
            },
        )

        # Format for when Completed = 0 (light yellow for incomplete)
        worksheet.conditional_format(
            f"A2:{last_col_letter}{last_row + 1}",  # Data range (excluding header)
            {
                "type": "formula",
                "criteria": f"=${completed_col_letter}2=0",  # Check if Completed column = 0
                "format": workbook.add_format({"bg_color": "#FFF8DC"}),  # Light yellow
            },
        )

    # Freeze the first row (header)
    worksheet.freeze_panes(1, 0)

    # Auto-filter for the entire data range
    worksheet.autofilter(0, 0, len(csv_rows), len(columns) - 1)

    workbook.close()
    print(
        f"XLSX file with validation and conditional formatting created: {output_path}"
    )


def load_xlsx_as_csv_rows(xlsx_path: str) -> list[dict]:
    """Load XLSX file and convert to CSV-like rows."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl not installed. Install with: pip install openpyxl")
        return []

    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    # Get headers from first row
    headers = []
    for cell in worksheet[1]:
        headers.append(cell.value or "")

    # Get data rows
    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(value) if value is not None else ""
        rows.append(row_dict)

    return rows


async def _judge_grammaticality_many(
    rows: list[dict],
    idxs: list[int],
    model_name: str,
    concurrency: int,
) -> list[dict]:
    class GrammarJudgment(BaseModel):
        grammatical: Literal["yes", "no", "partial"]
        explanation: Optional[str] = None

    schema = GrammarJudgment.model_json_schema()
    from openai import AsyncOpenAI

    base = (
        os.getenv("OPENAI_API_BASE")
        or os.getenv("SGLANG_API_BASE")
        or "http://localhost:30000/v1"
    )
    key = os.getenv("OPENAI_API_KEY") or "EMPTY"
    client = AsyncOpenAI(api_key=key, base_url=base)
    sem = asyncio.Semaphore(concurrency)

    async def one(text: str) -> dict:
        instr = (
            "Judge ONLY the grammatical acceptability of the MARKED modal verb in the English utterance. "
            "The modal is enclosed in asterisks, e.g., *must*. "
            "Consider the marked modal's form, placement, agreement, and clause context. "
            "Ignore typos, misspellings, and any errors outside the marked span. "
            "Return strict JSON with fields: grammatical ('yes'|'no'|'partial') and optional explanation."
        )

        def _strip_code_fences(s: str) -> str:
            s = s.strip()
            s = re.sub(r"^\s*```(?:json)?\s*", "", s, flags=re.IGNORECASE)
            s = re.sub(r"\s*```\s*$", "", s)
            return s

        def _parse_json(s: str) -> dict | None:
            s1 = _strip_code_fences(s)
            try:
                return json.loads(s1)
            except Exception:
                m = re.search(r"\{.*\}", s, flags=re.DOTALL)
                if m:
                    try:
                        return json.loads(m.group(0))
                    except Exception:
                        return None
                return None

        async with sem:
            try:
                resp = await client.responses.create(
                    model=model_name,
                    instructions=instr,
                    input=text,
                    text={
                        "format": {
                            "type": "json_schema",
                            "json_schema": {
                                "name": "GrammarJudgment",
                                "schema": schema,
                            },
                        },
                    },
                    temperature=0.0,
                )
                out = (getattr(resp, "output_text", "") or "").strip()
                data = _parse_json(out)
                if data is None:
                    logging.error(f"JSON parse failed: {out} for text: {text}")
                if not (
                    isinstance(data, dict)
                    and data.get("grammatical") in ("yes", "no", "partial")
                ):
                    resp2 = await client.responses.create(
                        model=model_name,
                        instructions=instr,
                        input=text,
                        text={
                            "format": {
                                "type": "json_schema",
                                "json_schema": {
                                    "name": "GrammarJudgment",
                                    "schema": schema,
                                },
                            },
                        },
                        temperature=0.1,
                    )
                    out2 = (getattr(resp2, "output_text", "") or "").strip()
                    data = _parse_json(out2)
                    if data is None:
                        logging.error(
                            f"JSON parse failed (retry): {out2} for text: {text}"
                        )
                if isinstance(data, dict) and data.get("grammatical") in (
                    "yes",
                    "no",
                    "partial",
                ):
                    return data
                lower = (out or "").lower()
                for opt in ("yes", "no", "partial"):
                    if opt in lower:
                        return {"grammatical": opt, "explanation": out}
                return {"grammatical": "yes", "explanation": ""}
            except Exception as e:
                logging.error(f"Judging request failed: {e} for text {text}")
                return {"grammatical": "yes", "explanation": f"fallback: {e}"}

    def _ensure_marked(row: dict) -> str:
        s = row.get("Marked_Sentence_English", "") or ""
        if "*" in s:
            return s
        orig = row.get("Original_Sentence", "") or ""
        mv = row.get("Tested_Modal", "") or ""
        if orig and mv:
            pat = r"\b" + re.escape(mv) + r"\b"
            marked = re.sub(
                pat,
                lambda m: f"*{m.group(0)}*",
                orig,
                count=1,
                flags=re.IGNORECASE,
            )
            return marked
        return s or orig

    texts = [_ensure_marked(rows[i]) for i in idxs]

    async def _one_with_index(i: int, text: str) -> tuple[int, dict]:
        return i, await one(text)

    tasks = [asyncio.create_task(_one_with_index(i, t)) for i, t in enumerate(texts)]
    results: list[dict] = [None] * len(texts)
    bar = tqdm(total=len(tasks), desc="Judging grammaticality", unit="row")
    for fut in asyncio.as_completed(tasks):
        i, res = await fut
        results[i] = res
        bar.update(1)
    bar.close()
    return results


async def _repair_grammaticality_many(
    rows: list[dict],
    idxs: list[int],
    model_name: str,
    concurrency: int,
) -> list[str]:
    from openai import AsyncOpenAI

    base = (
        os.getenv("OPENAI_API_BASE")
        or os.getenv("SGLANG_API_BASE")
        or "http://localhost:30000/v1"
    )
    key = os.getenv("OPENAI_API_KEY") or "EMPTY"
    client = AsyncOpenAI(api_key=key, base_url=base)
    sem = asyncio.Semaphore(concurrency)

    def _ensure_marked(row: dict) -> str:
        s = row.get("Marked_Sentence_English", "") or ""
        if "*" in s:
            return s
        orig = row.get("Original_Sentence", "") or ""
        mv = row.get("Tested_Modal", "") or ""
        if orig and mv:
            pat = r"\b" + re.escape(mv) + r"\b"
            return re.sub(
                pat, lambda m: f"*{m.group(0)}*", orig, count=1, flags=re.IGNORECASE
            )
        return s or orig

    def _build_repair_instructions(row: dict) -> str:
        strat = (row.get("Transformation_Strategy") or "original").lower()
        tested = row.get("Tested_Modal", "") or ""
        src = row.get("Source_Modal", "") or ""
        if strat == "remove_modality":
            return (
                "Rewrite the utterance by removing the marked modal while preserving grammar, polarity, meaning, and clause boundaries. "
                "Do not add modal adverbs; fix negation scope; preserve lead adverbials and expletives. Return exactly one sentence."
            )
        return (
            f"You will receive an English utterance with the tested modal marked between asterisks (e.g., *must*). "
            f"Rewrite ONLY the substring inside the first pair of asterisks to use the tested modal '{tested}' (not '{src}') in a grammatical way. "
            "You MAY adjust up to 5 tokens immediately to the right of the stars to fix tense/aspect (e.g., add 'have/been/going to'), "
            "but copy all other text EXACTLY as-is: do not insert, delete, or reorder outside the marked span and the small right-side window. "
            "Keep exactly one pair of asterisks in the output around the rewritten modal phrase. Return exactly one sentence."
        )

    def _split_star(s: str) -> tuple[str, str, str] | None:
        parts = s.split("*")
        if len(parts) < 3:
            return None
        return parts[0], parts[1], "*".join(parts[2:])

    _TOK_RE = re.compile(r"\w+|[^\w\s]")

    def _tokens(s: str) -> list[str]:
        return [m.group(0) for m in _TOK_RE.finditer(s)]

    def _outside_change_ok(
        orig_marked: str,
        cand_marked: str,
        k_left: int = 2,
        k_right: int = 5,
        max_del_ratio: float = 0.25,
        max_edit_ratio: float = 0.35,
    ) -> bool:
        o = _split_star(orig_marked)
        c = _split_star(cand_marked)
        if not o or not c:
            return False
        o_pre, _, o_suf = o
        c_pre, _, c_suf = c
        o_pre_t = _tokens(o_pre)
        c_pre_t = _tokens(c_pre)
        o_suf_t = _tokens(o_suf)
        c_suf_t = _tokens(c_suf)
        # Trim safety windows: allow changes in last k_left tokens of prefix and first k_right tokens of suffix
        o_pre_trim = o_pre_t[:-k_left] if len(o_pre_t) > k_left else []
        c_pre_trim = c_pre_t[:-k_left] if len(c_pre_t) > k_left else []
        o_suf_trim = o_suf_t[k_right:] if len(o_suf_t) > k_right else []
        c_suf_trim = c_suf_t[k_right:] if len(c_suf_t) > k_right else []
        # Combine trimmed outsides
        o_trim = o_pre_trim + o_suf_trim
        c_trim = c_pre_trim + c_suf_trim
        # Diff stats
        import difflib

        sm = difflib.SequenceMatcher(a=o_trim, b=c_trim)
        deletions = insertions = replacements = 0
        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == "delete":
                deletions += i2 - i1
            elif tag == "insert":
                insertions += j2 - j1
            elif tag == "replace":
                replacements += max(i2 - i1, j2 - j1)
        base = max(1, len(o_trim))
        del_ratio = deletions / base
        edit_ratio = (insertions + replacements) / base
        # Also bound overall length change ratio
        len_ratio = abs(len(c_trim) - len(o_trim)) / base
        return (
            (del_ratio <= max_del_ratio)
            and (edit_ratio <= max_edit_ratio)
            and (len_ratio <= max_edit_ratio)
        )

    async def _one(row: dict) -> str:
        strat = (row.get("Transformation_Strategy") or "original").lower()
        if strat == "remove_modality":
            from modality_llm.augmentation import remove_modality_transform_api_async

            async with sem:
                return await remove_modality_transform_api_async(
                    row.get("Original_Sentence", "") or "",
                    row.get("Source_Modal", "") or "",
                    None,
                    model_name=model_name,
                    base_url=base,
                )
        marked = _ensure_marked(row)
        instr = _build_repair_instructions(row)
        async with sem:
            try:
                resp = await client.responses.create(
                    model=model_name,
                    instructions=instr,
                    input=marked,
                    temperature=0.0,
                )
                out = (
                    (getattr(resp, "output_text", "") or "")
                    .strip()
                    .strip('"')
                    .strip("'")
                )
                valid = bool(out)
                if valid:
                    valid = _outside_change_ok(marked, out)
                if not valid or out == marked:
                    instr2 = (
                        instr
                        + " Your previous answer changed text outside the allowed window. Only change inside the stars and up to 5 tokens immediately to the right; keep all other text exactly unchanged."
                    )
                    resp2 = await client.responses.create(
                        model=model_name,
                        instructions=instr2,
                        input=marked,
                        temperature=0.1,
                    )
                    out2 = (
                        (getattr(resp2, "output_text", "") or "")
                        .strip()
                        .strip('"')
                        .strip("'")
                    )
                    if out2 and _outside_change_ok(marked, out2):
                        out = out2
                        valid = True
                if not valid:
                    return ""  # reject repair; caller keeps original text
                # ensure modal marking exists
                if "*" not in out:
                    tested = row.get("Tested_Modal", "") or ""
                    if tested:
                        pat = r"\b" + re.escape(tested) + r"\b"
                        out = re.sub(
                            pat,
                            lambda m: f"*{m.group(0)}*",
                            out,
                            count=1,
                            flags=re.IGNORECASE,
                        )
                return out
            except Exception as e:
                logging.error(f"Repair request failed: {e}")
                return ""

    texts = [rows[i] for i in idxs]

    async def _one_with_index(i: int, row: dict) -> tuple[int, str]:
        return i, await _one(row)

    tasks = [asyncio.create_task(_one_with_index(i, r)) for i, r in enumerate(texts)]
    results: list[str] = [""] * len(tasks)
    bar = tqdm(total=len(tasks), desc="Repairing grammaticality", unit="row")
    for fut in asyncio.as_completed(tasks):
        i, res = await fut
        results[i] = res
        bar.update(1)
    bar.close()
    return results


def generate_grammar_examples_for_annotation(
    modal_data_path: str,
    output_csv_path: str,
    include_alternatives: bool = True,
    require_consensus: str | None = None,
    output_format: str = "csv",
    existing_csv_path: str | None = None,
    freeze_completed: bool = True,
    mark_diff: bool = True,
    removal_backend: str = "spacy",
    removal_model: str = "openai/gpt-oss-20b",
    removal_concurrency: int = 8,
    judge_grammaticality: bool = False,
    judge_model: str = "openai/gpt-oss-20b",
    judge_concurrency: int = 8,
    judge_overwrite: bool = False,
    judge_only_augmented: bool = True,
) -> list[Example]:
    modal_path = Path(modal_data_path)
    output_path = Path(output_csv_path)

    if not modal_path.exists():
        print(f"Error: File not found at {modal_path}")
        return []

    output_path.parent.mkdir(parents=True, exist_ok=True)

    from modality_llm.schema import ModalExample
    from modality_llm.utils import unanimous_examples

    modal_data = load_jsonl_models(str(modal_path), ModalExample)
    # filter if requested
    modal_data = unanimous_examples(modal_data, require_consensus)
    if not modal_data:
        print("Error: No data loaded from modal dataset.")
        return []

    ENGLISH_MODALS: List[str] = [
        "can",
        "could",
        "may",
        "might",
        "must",
        "shall",
        "should",
        "will",
        "would",
        "ought to",
    ]

    print(
        f"Generating grammar examples for annotation from {len(modal_data)} modal sentences..."
    )
    print(f"Output will be saved to: {output_path}")
    if include_alternatives:
        print("Including alternative modal verb substitutions.")

    examples: list[Example] = []
    csv_rows = []
    removal_jobs: list[dict] = []

    def predict_grammaticality(transformation_strategy: str) -> str:
        """Predict if transformation likely preserves grammaticality."""
        ungrammatical_patterns = [
            "ability_to_denial",  # can → cannot (often ungrammatical)
            "necessity_to_denial",  # must → need not (often ungrammatical)
            "double_modal",  # might can (always ungrammatical)
            "insert_to",  # can to go (always ungrammatical)
        ]

        partial_patterns = [
            "necessity_to_possibility",  # must → may (weakens meaning significantly)
            "advice_to_possibility",  # should → can (changes obligation to permission)
            "prediction_to_possibility",  # will → might (major certainty change)
            "conditional_to_possibility",  # would → might (context-dependent)
        ]

        if transformation_strategy in ungrammatical_patterns:
            return "no"
        elif transformation_strategy in partial_patterns:
            return "partial"
        else:
            return "yes"

    for entry in modal_data:
        original_utt = entry.utt
        original_mv = entry.mv.lower()

        pattern = r"\b" + re.escape(original_mv) + r"\b"
        match = re.search(pattern, original_utt, re.IGNORECASE)

        if not match:
            print(
                f"Warning: Could not find modal '{original_mv}' in utterance: {original_utt}"
            )
            continue

        start, end = match.span()
        found_modal = original_utt[start:end]
        eid = hashlib.sha256(original_utt.encode("utf-8")).hexdigest()

        # Create Example for original, using sha256(utt) as eid
        human_annotations = entry.annotations
        expected_categories = entry.res
        original_example = Example(
            eid=eid,
            english=f"{original_utt[:start]}*{found_modal}*{original_utt[end:]}",
            japanese=None,
            grammatical=GrammarLabel.yes,
            english_target=original_mv,
            japanese_target=None,
            human_annotations=human_annotations,
            expected_categories=expected_categories,
        )
        examples.append(original_example)
        csv_rows.append(
            {
                "ID": f"{eid}_orig",
                "EID": eid,
                "Original_Sentence": original_utt,
                "Modal_Verb": original_mv,
                "Marked_Sentence_English": original_example.english,
                "Marked_Sentence_Japanese": "",
                "Source_Modal": original_mv,
                "Tested_Modal": original_mv,
                "Transformation_Strategy": "original",
                "Transformation_Category": get_transformation_category("original"),
                "Palmer_Annotations": json.dumps(human_annotations.get("palmer", []))
                if human_annotations
                else "[]",
                "Quirk_Annotations": json.dumps(human_annotations.get("quirk", []))
                if human_annotations
                else "[]",
                "Palmer_Expected": expected_categories.get("palmer", ["unknown"])[0]
                if expected_categories and expected_categories.get("palmer")
                else "unknown",
                "Quirk_Expected": expected_categories.get("quirk", ["unknown"])[0]
                if expected_categories and expected_categories.get("quirk")
                else "unknown",
                "Grammaticality_Expected": "yes",
                "Completed": 1,
                "Annotation_Notes": f"Original sentence from dataset. Modal: '{original_mv}'. Verified and complete.",
            }
        )

        if include_alternatives:
            # Mirror original modal capitalization for alternative insertions
            def _match_case(word: str, template: str) -> str:
                if not template:
                    return word
                if template[0].isupper():
                    return (
                        "Ought to"
                        if word.lower() == "ought to"
                        else (word[0].upper() + word[1:])
                    )
                return word

            for alt_mv in ENGLISH_MODALS:
                if alt_mv.lower() == original_mv.lower():
                    continue

                # Use the actual augmentation system to determine strategy
                transformation_strategy = (
                    determine_transformation_strategy_from_augmentation(
                        original_example, alt_mv
                    )
                )

                alt_eid = f"{eid}_alt_{alt_mv.replace(' ', '_')}"
                alt_mv_cased = _match_case(alt_mv, found_modal)
                alt_example = Example(
                    eid=alt_eid,
                    english=f"{original_utt[:start]}*{alt_mv_cased}*{original_utt[end:]}",
                    japanese=None,
                    grammatical=GrammarLabel.yes,
                    english_target=alt_mv,
                    japanese_target=None,
                    human_annotations=human_annotations,
                    expected_categories=expected_categories,
                )
                examples.append(alt_example)

                # decide whether we assume grammaticality for this strategy (remove_modality is assumed yes)
                gr_expected = (
                    "yes"
                    if transformation_strategy == "remove_modality"
                    else predict_grammaticality(transformation_strategy)
                )
                if transformation_strategy == "remove_modality":
                    ann_note = (
                        f"Modal substitution: '{original_mv}' → '{alt_mv}'. Strategy: {transformation_strategy}. "
                        "ASSUMED GRAMMATICAL (not verified). Please classify modal categories."
                    )
                else:
                    ann_note = (
                        f"Modal substitution: '{original_mv}' → '{alt_mv}'. Strategy: {transformation_strategy}. "
                        "Please verify grammaticality and classify modal categories."
                    )

                csv_rows.append(
                    {
                        "ID": alt_eid,
                        "EID": eid,
                        "Original_Sentence": original_utt,
                        "Modal_Verb": alt_mv,
                        "Marked_Sentence_English": alt_example.english,
                        "Marked_Sentence_Japanese": "",
                        "Source_Modal": original_mv,
                        "Tested_Modal": alt_mv,
                        "Transformation_Strategy": transformation_strategy,
                        "Transformation_Category": get_transformation_category(
                            transformation_strategy
                        ),
                        "Palmer_Annotations": json.dumps(
                            human_annotations.get("palmer", [])
                        )
                        if human_annotations
                        else "[]",
                        "Quirk_Annotations": json.dumps(
                            human_annotations.get("quirk", [])
                        )
                        if human_annotations
                        else "[]",
                        "Palmer_Expected": "unknown",
                        "Quirk_Expected": "unknown",
                        "Grammaticality_Expected": gr_expected,
                        "Completed": 0,
                        "Annotation_Notes": ann_note,
                    }
                )

            # Collect a removal job to be run later (API concurrently or spaCy synchronously).
            removal_jobs.append(
                {
                    "eid": eid,
                    "original_utt": original_utt,
                    "original_mv": original_mv,
                    "start": start,
                    "end": end,
                    "found_modal": found_modal,
                    "human_annotations": human_annotations,
                    "expected_categories": expected_categories,
                }
            )

    if include_alternatives and removal_jobs:
        if removal_backend.lower() == "api":
            from modality_llm.augmentation import remove_modality_transform_api_async

            base = os.getenv("OPENAI_API_BASE") or os.getenv("SGLANG_API_BASE") or None
            print(
                f"Using API removal backend (model={removal_model}, base={base or 'env-default'}, concurrency={removal_concurrency})"
            )

            async def _run_all():
                sem = asyncio.Semaphore(removal_concurrency)

                async def _one(job):
                    async with sem:
                        return await remove_modality_transform_api_async(
                            job["original_utt"],
                            job["original_mv"],
                            None,
                            model_name=removal_model,
                            base_url=base,
                        )

                return await asyncio.gather(*[_one(j) for j in removal_jobs])

            removed_list = asyncio.run(_run_all())
        else:
            from modality_llm.augmentation import remove_modality_transform

            removed_list = [
                remove_modality_transform(j["original_utt"], j["original_mv"], None)
                for j in removal_jobs
            ]

        # Append removal rows and examples
        for job, removed_sentence in zip(removal_jobs, removed_list):
            removed_marked = mark_changed_tokens(
                job["original_utt"],
                removed_sentence,
                focus_span=(job["start"], job["end"]),
            )
            rem_alt_mv_label = "(removed)"
            rem_alt_eid = f"{job['eid']}_alt_removed"
            rem_alt_example = Example(
                eid=rem_alt_eid,
                english=removed_marked,
                japanese=None,
                grammatical=GrammarLabel.yes,
                english_target=rem_alt_mv_label,
                japanese_target=None,
                human_annotations=job["human_annotations"],
                expected_categories=job["expected_categories"],
            )
            examples.append(rem_alt_example)
            csv_rows.append(
                {
                    "ID": rem_alt_eid,
                    "EID": job["eid"],
                    "Original_Sentence": job["original_utt"],
                    "Modal_Verb": rem_alt_mv_label,
                    "Marked_Sentence_English": rem_alt_example.english,
                    "Marked_Sentence_Japanese": "",
                    "Source_Modal": job["original_mv"],
                    "Tested_Modal": rem_alt_mv_label,
                    "Transformation_Strategy": "remove_modality",
                    "Transformation_Category": get_transformation_category(
                        "remove_modality"
                    ),
                    "Palmer_Annotations": json.dumps(
                        job["human_annotations"].get("palmer", [])
                    )
                    if job["human_annotations"]
                    else "[]",
                    "Quirk_Annotations": json.dumps(
                        job["human_annotations"].get("quirk", [])
                    )
                    if job["human_annotations"]
                    else "[]",
                    "Palmer_Expected": "unknown",
                    "Quirk_Expected": "unknown",
                    "Grammaticality_Expected": "yes",
                    "Completed": 0,
                    "Removal_Backend": removal_backend,
                    "Annotation_Notes": (
                        f"Modal substitution: '{job['original_mv']}' → (removed). Strategy: remove_modality. "
                        "ASSUMED GRAMMATICAL (not verified). Please classify modal categories."
                    ),
                }
            )

    if judge_grammaticality:
        to_judge = []
        for i, row in enumerate(csv_rows):
            if row.get("Transformation_Strategy") == "remove_modality":
                # removal rows are assumed grammatical and may not contain marked insertions
                row["Grammaticality_Expected"] = "yes"
                continue
            if not judge_overwrite and str(row.get("Completed", "0")) == "1":
                continue
            if (
                judge_only_augmented
                and row.get("Transformation_Strategy") == "original"
            ):
                continue
            to_judge.append(i)
        if to_judge:
            results = asyncio.run(
                _judge_grammaticality_many(
                    csv_rows, to_judge, judge_model, judge_concurrency
                )
            )
            for idx, res in zip(to_judge, results):
                csv_rows[idx]["Grammaticality_Expected"] = res.get("grammatical", "yes")
                if res.get("explanation"):
                    csv_rows[idx]["Grammaticality_Explanation"] = res["explanation"]

    try:
        if output_format.lower() == "xlsx":
            generate_grammar_examples_xlsx(csv_rows, output_csv_path)
        else:
            import polars as pl

            # Merge / diff support: carry forward completed flags and labels from an existing CSV
            if existing_csv_path:
                new_df = pl.DataFrame(csv_rows)
                # Stable key for rows
                new_df = new_df.with_columns(
                    pl.concat_str(
                        [
                            pl.col("EID"),
                            pl.lit("|"),
                            pl.col("Tested_Modal"),
                            pl.lit("|"),
                            pl.col("Transformation_Strategy"),
                        ]
                    ).alias("KEY")
                )
                old_df = pl.read_csv(existing_csv_path)
                if "Transformation_Strategy" not in old_df.columns:
                    old_df = old_df.with_columns(
                        pl.lit("original").alias("Transformation_Strategy")
                    )
                old_df = old_df.with_columns(
                    pl.concat_str(
                        [
                            pl.col("EID"),
                            pl.lit("|"),
                            pl.col("Tested_Modal"),
                            pl.lit("|"),
                            pl.col("Transformation_Strategy"),
                        ]
                    ).alias("KEY")
                )
                carry = old_df.select(
                    [
                        "KEY",
                        "Completed",
                        "Palmer_Expected",
                        "Quirk_Expected",
                        "Grammaticality_Expected",
                        "Marked_Sentence_English",
                    ]
                ).rename(
                    {
                        "Completed": "Old_Completed",
                        "Palmer_Expected": "Old_Palmer_Expected",
                        "Quirk_Expected": "Old_Quirk_Expected",
                        "Grammaticality_Expected": "Old_Grammaticality_Expected",
                        "Marked_Sentence_English": "Old_Text",
                    }
                )
                merged = new_df.join(carry, on="KEY", how="left")
                # Coalesce labels/completed
                merged = merged.with_columns(
                    [
                        pl.coalesce(
                            [pl.col("Old_Palmer_Expected"), pl.col("Palmer_Expected")]
                        ).alias("Palmer_Expected"),
                        pl.coalesce(
                            [pl.col("Old_Quirk_Expected"), pl.col("Quirk_Expected")]
                        ).alias("Quirk_Expected"),
                        pl.coalesce(
                            [
                                pl.col("Old_Grammaticality_Expected"),
                                pl.col("Grammaticality_Expected"),
                            ]
                        ).alias("Grammaticality_Expected"),
                        pl.coalesce(
                            [
                                pl.col("Old_Completed").cast(pl.Utf8),
                                pl.col("Completed").cast(pl.Utf8),
                            ]
                        ).alias("Completed"),
                    ]
                )
                # Detect changes
                merged = merged.with_columns(
                    (pl.col("Marked_Sentence_English") != pl.col("Old_Text"))
                    .fill_null(False)
                    .alias("Changed")
                )
                # Optionally freeze text for completed rows
                if freeze_completed:
                    merged = merged.with_columns(
                        pl.when((pl.col("Old_Completed") == "1") & pl.col("Changed"))
                        .then(pl.col("Old_Text"))
                        .otherwise(pl.col("Marked_Sentence_English"))
                        .alias("Marked_Sentence_English")
                    )
                # Optional marked diff column using utils.mark_changed_tokens
                if mark_diff:

                    def _mk_diff(
                        orig: str | None, new: str | None, orig_utt: str, mv: str
                    ) -> str:
                        if not orig or not new or orig == new:
                            return new or ""
                        # anchor on the original modal span if we can find it
                        m = re.search(
                            r"\b" + re.escape(mv) + r"\b", orig_utt, flags=re.IGNORECASE
                        )
                        if m:
                            return mark_changed_tokens(
                                orig, new, focus_span=(m.span(0)[0], m.span(0)[1])
                            )
                        return mark_changed_tokens(orig, new)

                    rows = merged.to_dicts()
                    for r in rows:
                        r["Marked_Diff"] = _mk_diff(
                            r.get("Old_Text"),
                            r.get("Marked_Sentence_English"),
                            r.get("Original_Sentence", ""),
                            r.get("Source_Modal", ""),
                        )
                    # Replace new_df with merged rows including Marked_Diff
                    new_df = pl.DataFrame(rows)
                else:
                    new_df = merged

                # Drop helper columns and use merged frame for output
                drop_cols = [
                    "KEY",
                    "Old_Completed",
                    "Old_Palmer_Expected",
                    "Old_Quirk_Expected",
                    "Old_Grammaticality_Expected",
                    "Old_Text",
                    "Changed",
                ]
                new_df = new_df.drop([c for c in drop_cols if c in new_df.columns])

                new_df.write_csv(str(output_path))
            else:
                pl.DataFrame(csv_rows).write_csv(str(output_path))
        print(
            f"Successfully generated {len(csv_rows)} examples for annotation in {output_path}"
        )
        return examples
    except Exception as e:
        print(f"Error writing {output_format.upper()} file: {e}")
        return []


def judge_grammaticality_in_csv(
    csv_path: str,
    output_csv_path: str,
    model_name: str = "openai/gpt-oss-20b",
    concurrency: int = 8,
    overwrite: bool = False,
    only_augmented: bool = True,
) -> None:
    import polars as pl

    from modality_llm.utils import load_csv

    rows = load_csv(csv_path)
    if not rows:
        print(f"Error: No rows loaded from {csv_path}")
        return

    # Always assume removal rows are grammatical and skip judging them
    for r in rows:
        if r.get("Transformation_Strategy") == "remove_modality":
            r["Grammaticality_Expected"] = "yes"

    idxs = []
    for i, r in enumerate(rows):
        if r.get("Transformation_Strategy") == "remove_modality":
            continue
        if not overwrite and str(r.get("Completed", "0")) == "1":
            continue
        if only_augmented and r.get("Transformation_Strategy") == "original":
            continue
        idxs.append(i)

    if not idxs:
        print("Nothing to judge (no eligible rows).")
        pl.DataFrame(rows).write_csv(output_csv_path)
        print(f"Wrote CSV to {output_csv_path}")
        return

    results = asyncio.run(
        _judge_grammaticality_many(rows, idxs, model_name, concurrency)
    )
    for i, res in zip(idxs, results):
        rows[i]["Grammaticality_Expected"] = res.get("grammatical", "yes")
        if res.get("explanation"):
            rows[i]["Grammaticality_Explanation"] = res["explanation"]

    pl.DataFrame(rows).write_csv(output_csv_path)
    print(f"Judged {len(idxs)} rows; wrote CSV to {output_csv_path}")


def repair_grammaticality_in_csv(
    csv_path: str,
    output_csv_path: str,
    model_name: str = "openai/gpt-oss-20b",
    concurrency: int = 8,
    only_augmented: bool = True,
    only_bad: bool = True,
    rejudge: bool = True,
) -> None:
    import polars as pl

    from modality_llm.utils import load_csv

    rows = load_csv(csv_path)
    if not rows:
        print(f"Error: No rows loaded from {csv_path}")
        return
    idxs: list[int] = []
    for i, r in enumerate(rows):
        strat = (r.get("Transformation_Strategy") or "").lower()
        if only_augmented and strat == "original":
            continue
        if strat == "remove_modality":
            r["Grammaticality_Expected"] = "yes"
            continue
        if only_bad and r.get("Grammaticality_Expected", "yes") not in (
            "no",
            "partial",
        ):
            continue
        idxs.append(i)
    if not idxs:
        print("Nothing to repair (no eligible rows).")
        pl.DataFrame(rows).write_csv(output_csv_path)
        print(f"Wrote CSV to {output_csv_path}")
        return
    repaired = asyncio.run(
        _repair_grammaticality_many(rows, idxs, model_name, concurrency)
    )
    for i, new_text in zip(idxs, repaired):
        if new_text:
            rows[i]["Marked_Sentence_English"] = new_text
            rows[i]["Repair_Status"] = "api"
    if rejudge:
        judged = asyncio.run(
            _judge_grammaticality_many(rows, idxs, model_name, concurrency)
        )
        for i, res in zip(idxs, judged):
            rows[i]["Grammaticality_Expected"] = res.get("grammatical", "yes")
            if res.get("explanation"):
                rows[i]["Grammaticality_Explanation"] = res["explanation"]
    pl.DataFrame(rows).write_csv(output_csv_path)
    print(f"Repaired {len(idxs)} rows; wrote CSV to {output_csv_path}")


def convert_annotated_csv_to_jsonl(
    csv_path: str,
    output_jsonl_path: str,
    completed_only: bool = False,
    jsonl_format: str = "minimal",
) -> None:
    """Convert annotated CSV/XLSX back to modal_verbs.jsonl format, or pass through existing JSONL."""
    csv_path_obj = Path(csv_path)
    output_path_obj = Path(output_jsonl_path)

    if not csv_path_obj.exists():
        print(f"Error: File not found at {csv_path}")
        return

    output_path_obj.parent.mkdir(parents=True, exist_ok=True)

    # Check if input is already a JSONL file
    if csv_path_obj.suffix.lower() == ".jsonl":
        print(
            f"Input is already JSONL format. Copying {csv_path} to {output_jsonl_path}"
        )
        try:
            import shutil

            shutil.copy2(csv_path_obj, output_path_obj)
            print(f"Successfully copied JSONL file to {output_path_obj}")
        except Exception as e:
            print(f"Error copying JSONL file: {e}")
        return

    # Load data based on file extension
    if csv_path_obj.suffix.lower() == ".xlsx":
        rows = load_xlsx_as_csv_rows(str(csv_path_obj))
    else:
        from modality_llm.utils import load_csv

        rows = load_csv(str(csv_path_obj))

    if not rows:
        print("Error: No data loaded from CSV/XLSX file.")
        return

    # Check if this looks like an annotated CSV (has our expected columns)
    if not rows or "ID" not in rows[0]:
        print(
            "Error: Input CSV/XLSX does not appear to be an annotated file with required columns."
        )
        print("Expected columns include: ID, EID, Palmer_Annotations, etc.")
        print("Use 'csv' mode to generate an annotated CSV/XLSX first.")
        return

    # Optionally filter for completed examples only
    if completed_only:
        original_count = len(rows)
        rows = [row for row in rows if str(row.get("Completed", "0")) == "1"]
        print(
            f"Filtered to {len(rows)} completed examples (from {original_count} total)"
        )

    modal_examples = []
    for row in rows:
        try:
            palmer_final = row.get("Palmer_Expected", "unknown")
            quirk_final = row.get("Quirk_Expected", "unknown")
            grammaticality = row.get("Grammaticality_Expected", "yes")

            palmer_annotations = json.loads(row.get("Palmer_Annotations", "[]"))
            quirk_annotations = json.loads(row.get("Quirk_Annotations", "[]"))

            if jsonl_format == "minimal":
                modal_example = {
                    "mv": row.get("Tested_Modal", ""),
                    "utt": row.get("Marked_Sentence_English", "").replace("*", ""),
                    "annotations": {
                        "palmer": palmer_annotations,
                        "quirk": quirk_annotations,
                    },
                    "res": {
                        "palmer": [palmer_final] if palmer_final != "unknown" else [],
                        "quirk": [quirk_final] if quirk_final != "unknown" else [],
                    },
                }
            else:
                modal_example = {
                    "mv": row.get("Tested_Modal", ""),
                    "utt": row.get("Marked_Sentence_English", "").replace("*", ""),
                    "transformation_strategy": row.get(
                        "Transformation_Strategy", "original"
                    ),
                    "source_eid": row.get("EID", ""),
                    "completed": int(row.get("Completed", "0")),
                    "grammatical": grammaticality.lower(),
                    "grammatical_binary": grammaticality.lower() in ["yes", "partial"],
                    "annotations": {
                        "palmer": palmer_annotations,
                        "quirk": quirk_annotations,
                    },
                    "res": {
                        "palmer": [palmer_final] if palmer_final != "unknown" else [],
                        "quirk": [quirk_final] if quirk_final != "unknown" else [],
                    },
                }
            modal_examples.append(modal_example)

        except (json.JSONDecodeError, KeyError) as e:
            print(f"Warning: Skipping malformed row: {e}")
            continue

    # Write to JSONL
    try:
        with open(output_path_obj, "w", encoding="utf-8") as f:
            for example in modal_examples:
                json.dump(example, f, ensure_ascii=False)
                f.write("\n")

        print(
            f"Successfully converted {len(modal_examples)} examples to {output_path_obj}"
        )

    except Exception as e:
        print(f"Error writing JSONL file: {e}")
